from PyQt6.QtWidgets import QMainWindow,QApplication,QPushButton,QWidget,QVBoxLayout,QHBoxLayout,QSpinBox,QLabel,QLineEdit,QGridLayout,QTableWidget,QFileDialog,QDialog,QTableWidgetItem,QMessageBox,QDialogButtonBox,QGraphicsOpacityEffect
from PyQt6.QtGui import QPixmap
from PyQt6 import QtGui,QtCore
import sys
from random import choice
from openpyxl import workbook
import json
import os.path
import time
class window(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI_form1()
        self.initUI_form2()
        self.init_variable()
        self.load_setting()

    def init_variable(self):
        self.masiv_prizes={}
        self.masiv_uchastnikov=[]
        self.prize_id=0

    def initUI_form1(self):
        self.setWindowTitle("Рандомайзер подарков")

        self.setting_layout=QHBoxLayout(self)
        self.bt_settings=QPushButton("Настройки",self)
        self.setting_layout.addStretch()
        self.setting_layout.addWidget(self.bt_settings)
        self.bt_settings.clicked.connect(self.form2_show)

        self.bt_reboot=QPushButton("Перезапуск",self)
        self.bt_reboot.hide()
        self.bt_reboot.clicked.connect(self.reboot)
        self.setting_layout.addWidget(self.bt_reboot)


        self.name_layout=QHBoxLayout()
        self.name_prizes=QLabel("Стартовый текст")
        self.name_prizes.setFont(QtGui.QFont("Times",20))
        self.name_layout.addStretch()
        self.name_layout.addWidget(self.name_prizes)
        self.name_layout.addStretch()

        self.image_layout=QHBoxLayout()
        self.image_label=QLabel()
        self.image=QPixmap("start_image.png")
        self.image_label.setPixmap(self.image)
        self.image_label.setScaledContents(True)
        self.image_label.setFixedSize(500,500)

        self.image_layout.addWidget(self.image_label)

        self.vin_count_layout=QHBoxLayout()
        self.vin_count_text=QLabel("Количество")
        self.vin_count_text.setFont(QtGui.QFont("Times",20))
        self.vin_count_text.hide()
        self.vin_count_layout.addStretch()
        self.vin_count_layout.addWidget(self.vin_count_text)
        self.vin_count_layout.addStretch()


        self.vin_text_layout=QHBoxLayout()
        self.vin_text_text=QLabel("Победители")
        self.vin_text_text.setFont(QtGui.QFont("Times",20))
        self.vin_text_text.hide()
        self.vin_text_layout.addStretch()
        self.vin_text_layout.addWidget(self.vin_text_text)
        self.vin_text_layout.addStretch()

        self.vin_layout=QHBoxLayout()
        # self.vin_layout=QGridLayout()
        self.vin_text=QLabel("")
        self.vin_text.setFont(QtGui.QFont("Times",20))
        self.vin_text.hide()
        self.vin_layout.addStretch()
        self.vin_layout.addWidget(self.vin_text)
        # self.vin_layout.addStretch()

        self.vin_sort_layout=QHBoxLayout()
        self.vin_sort_text=QLabel("")
        self.vin_sort_text.setFont(QtGui.QFont("Times",20))
        self.vin_sort_text.hide()
        self.vin_sort_layout.addStretch()
        self.vin_sort_layout.addWidget(self.vin_sort_text)
        self.vin_sort_layout.addStretch()






        self.bt_layout=QHBoxLayout()
        self.bt_start=QPushButton("Старт")
        self.bt_start.setFont(QtGui.QFont("Times",20))
        self.bt_start.clicked.connect(self.start)

        self.bt_next=QPushButton("Следующий приз")
        self.bt_next.setFont(QtGui.QFont("Times",20))
        self.bt_next.clicked.connect(self.next_prize)
        self.bt_next.hide()

        self.bt_back=QPushButton("Предыдущий приз")
        self.bt_back.setFont(QtGui.QFont("Times",20))
        self.bt_back.clicked.connect(self.back_prize)
        self.bt_back.hide()



        self.bt_layout.addWidget(self.bt_start)
        self.bt_layout.addWidget(self.bt_back)
        self.bt_layout.addWidget(self.bt_next)





        self.strech_layout=QHBoxLayout()
        self.label_test=QLabel()
        self.strech_layout.addStretch()
        self.strech_layout.addWidget(self.label_test)

        self.main_layout=QVBoxLayout(self)
        self.main_layout.addLayout(self.setting_layout)

        self.main_layout.addLayout(self.name_layout)
        self.main_layout.addStretch()
        self.main_layout.addLayout(self.image_layout)
        self.main_layout.addLayout(self.vin_count_layout)
        self.main_layout.addLayout(self.vin_text_layout)
        self.main_layout.addLayout(self.vin_layout)
        self.main_layout.addLayout(self.vin_sort_layout)
        self.main_layout.addStretch()
        # self.main_layout.addLayout(self.strech_layout)
        self.main_layout.addLayout(self.bt_layout)

        central_widget=QWidget()
        central_widget.setLayout(self.main_layout)


        self.setCentralWidget(central_widget)

        self.showMaximized()

    def reboot(self):
        self.initUI_form1()
        self.initUI_form2()
        self.init_variable()
        self.load_setting()

    def initUI_form2(self):
        self.form=QDialog()
        self.form.setModal(True)


        # self.range_ot=QSpinBox()
        # self.range_ot.setMinimum(1)
        # self.range_ot.setMaximum(10000000)
        self.range_do=QSpinBox()
        self.range_do.setMinimum(2)
        self.range_do.setMaximum(10000000)
        self.label_range=QLabel()

        self.label_range.setText("Количество участников")
        # self.label_range_ot=QLabel()
        # self.label_range_do=QLabel()
        # self.label_range_ot.setText("От")
        # self.label_range_do.setText("До")

        range_layout_text=QHBoxLayout()
        range_layout_text.addStretch()
        range_layout_text.addWidget(self.label_range)
        range_layout_text.addStretch()
        range_layout=QHBoxLayout()
        # range_layout.addWidget(self.label_range_ot)
        # range_layout.addWidget(self.range_ot)
        # range_layout.addWidget(self.label_range_do)
        range_layout.addWidget(self.range_do)



        self.table_data=QTableWidget()
        self.table_data.setColumnCount(3)
        self.table_data.setHorizontalHeaderLabels(["Название","Путь","Количество"])
        self.table_data.horizontalHeader().setStretchLastSection(True)
        # self.table_data.resizeColumnsToContents()

        self.prizes_layout=QHBoxLayout()
        self.prizes_name=QLabel()
        self.prizes_name.setText("Название")
        self.prizes_name_edit=QLineEdit()

        self.prizes_put=QLabel()
        self.prizes_put.setText("Путь")
        self.prizes_put_edit=QLineEdit()
        self.bt_open_file=QPushButton()
        self.bt_open_file.setText("...")
        self.bt_open_file.clicked.connect(self.open_file)
        self.prizes_count=QLabel()
        self.prizes_count.setText("Количество")
        self.prizes_count_edit=QSpinBox()
        self.prizes_count_edit.setMinimum(1)
        self.prizes_count_edit.setMaximum(10000000)
        self.prizes_count_edit.setFixedWidth(100)

        # self.bt_delete_layout=QPushButton()
        # self.bt_delete_layout.setText("X")
        # self.bt_delete_layout.clicked.connect(self.delete_prizes)

        self.prizes_layout.addWidget(self.prizes_name)
        self.prizes_layout.addWidget(self.prizes_name_edit)
        self.prizes_layout.addWidget(self.prizes_put)
        self.prizes_layout.addWidget(self.prizes_put_edit)
        self.prizes_layout.addWidget(self.bt_open_file)
        self.prizes_layout.addWidget(self.prizes_count)
        self.prizes_layout.addWidget(self.prizes_count_edit)
        # self.prizes_layout.addWidget(self.bt_delete_layout)

        self.bt_add_prize=QPushButton()
        self.bt_add_prize.setText("Добавить призы")
        self.bt_add_prize.clicked.connect(self.add_prize)

        self.bt_delete_prize=QPushButton()
        self.bt_delete_prize.setText("Удалить")
        self.bt_delete_prize.clicked.connect(self.delete_data_table)

        self.upravlenie_bt=QHBoxLayout()
        self.upravlenie_bt.addWidget(self.bt_add_prize)
        self.upravlenie_bt.addWidget(self.bt_delete_prize)

        self.bt_save_prize=QPushButton()
        self.bt_save_prize.setText("Сохранить")
        self.bt_save_prize.clicked.connect(self.save)



        self.mail_text=QLabel()
        self.mail_text.hide()

        self.main_layout_form2=QGridLayout()
        self.main_layout_form2.addLayout(range_layout_text,0,0)
        self.main_layout_form2.addLayout(range_layout,1,0)
        self.main_layout_form2.addWidget(self.table_data,2,0)
        self.main_layout_form2.addLayout(self.prizes_layout,3,0)
        self.main_layout_form2.addLayout(self.upravlenie_bt,4,0)
        self.main_layout_form2.addWidget(self.bt_save_prize)
        self.main_layout_form2.addWidget(self.mail_text)

        central_widget=QWidget()
        central_widget.setLayout(self.main_layout_form2)
        self.form.setLayout(self.main_layout_form2)
        # self.form.setCentralWidget(central_widget)

    def start_animation(self):
        self.animate_numbers()

    def animate_numbers(self):
        if self.current_index < len(self.masiv_prizes["prize_data"][self.prize_id]['vin']):
           
            self.text[self.current_index].setText(str(self.masiv_prizes["prize_data"][self.prize_id]['vin'][self.current_index]))
            # self.text[self.current_index].setStyleSheet("background-color: lightgreen")
            self.current_index += 1
            QtCore.QTimer.singleShot(1000, self.animate_numbers)  # Вызываем метод снова через 500 миллисекунд

        else:
            self.vin_sort_text.show()

    def start(self):

        self.masiv_uchastnikov=[i for i in range(1,self.range_do.value()+1)]
        if "prize_data" in  self.masiv_prizes.keys():
            for masiv_prizes_id in range (len(self.masiv_prizes["prize_data"])):
                prize=self.masiv_prizes["prize_data"][masiv_prizes_id]
                prize_vin=[]
                for cont_prize in range(self.masiv_prizes["prize_data"][masiv_prizes_id]['count']):
                    vin=choice(self.masiv_uchastnikov)
                    prize_vin.append(vin)
                    self.masiv_uchastnikov.remove(vin)
                self.masiv_prizes["prize_data"][masiv_prizes_id]["vin"]=prize_vin
            # print(self.masiv_prizes)
            # print(self.masiv_uchastnikov)

            self.name_prizes.setText(self.masiv_prizes["prize_data"][0]["name"])
            image=QPixmap(self.masiv_prizes["prize_data"][0]["image"])
            self.image_label.setPixmap(image)
            self.vin_count_text.setText(f"Количество победителей {self.masiv_prizes["prize_data"][0]["count"]}")
            self.vin_count_text.show()

            vin_str=""

            self.text=[]
            count=0
            self.current_index=0
            for i in self.masiv_prizes["prize_data"][0]['vin']:
                self.text.append(QLabel(""))

                self.text[count].setFont(QtGui.QFont("Times",20))
                self.vin_layout.addWidget(self.text[count])


                # QApplication.processEvents()
                # self.unfade(self.text[i])


                count+=1
                # vin_str+=str(i)+" "
            self.start_animation()
            vin_sort_str=""
            self.vin_layout.addStretch()



            vin=self.masiv_prizes["prize_data"][0]['vin'].copy()
            vin.sort()

            for i in vin:
                vin_sort_str+=str(i)+" "
            self.vin_sort_text.setText(vin_sort_str)

            self.vin_text.setText(vin_str)

            self.vin_text_text.show()
            # self.vin_text.show()
            # self.unfade(self.vin_text)
            # self.vin_sort_text.show()
            self.bt_start.hide()
            if self.masiv_prizes["row_count"]>1:
                self.bt_next.show()
            self.prize_id=self.masiv_prizes["prize_data"][0]['id']


            self.wb=workbook.Workbook()
            self.ws=self.wb.active
            self.ws["A1"]="ID"
            self.ws["B1"]="Название приза"
            self.ws["C1"]="Количество призов"
            self.ws["D1"]="Победители"
            for i in range(len(self.masiv_prizes["prize_data"])):
                a="A"+str(i+2)
                b="B"+str(i+2)
                c="C"+str(i+2)
                d="D"+str(i+2)

                self.ws[a]=self.masiv_prizes["prize_data"][i]["id"]
                self.ws[b]=self.masiv_prizes["prize_data"][i]["name"]
                self.ws[c]=self.masiv_prizes["prize_data"][i]["count"]
                self.ws[d]=",".join(str(x) for x in self.masiv_prizes["prize_data"][i]["vin"])
            self.wb.save("vin.xlsx")

            self.bt_settings.hide()
            self.bt_reboot.show()
        else:
            msg = QMessageBox(
                parent=self,
                text="Нету призов",
                icon=QMessageBox.Icon.Warning

            )
            msg.setWindowTitle("Window")
            msg.exec()




    def next_prize(self):
        self.vin_sort_text.hide()
        self.prize_id+=1
        # print(self.masiv_prizes)
        self.bt_back.show()
        self.name_prizes.setText(self.masiv_prizes["prize_data"][self.prize_id]["name"])
        image=QPixmap(self.masiv_prizes["prize_data"][self.prize_id]["image"])
        self.image_label.setPixmap(image)
        self.vin_count_text.setText(f"Количество победителей {self.masiv_prizes["prize_data"][self.prize_id]["count"]}")
        self.vin_count_text.show()


        stretch_item=self.vin_layout.takeAt(self.vin_layout.count()-1)

        if stretch_item:
            del stretch_item


        for i in range(len(self.text)):
            self.vin_layout.removeWidget(self.text[i])
            self.text[i].setParent(None)


        vin_str=""
        self.text=[]
        count=0
        self.current_index=0
        for i in self.masiv_prizes["prize_data"][self.prize_id]['vin']:
            self.text.append(QLabel(""))

            self.text[count].setFont(QtGui.QFont("Times",20))
            self.vin_layout.addWidget(self.text[count])


            # QApplication.processEvents()
            # self.unfade(self.text[i])


            count+=1
            # vin_str+=str(i)+" "
        self.start_animation()
        self.vin_layout.addStretch()

        vin_sort_str=""
        vin=self.masiv_prizes["prize_data"][self.prize_id]['vin'].copy()
        vin.sort()

        for i in vin:
            vin_sort_str+=str(i)+" "

        self.vin_sort_text.setText(vin_sort_str)

        # self.vin_text.setText(vin_str)
        self.vin_text_text.show()
        # self.vin_text.show()
        self.bt_start.hide()
        self.bt_next.show()
        self.prize_id=self.masiv_prizes["prize_data"][self.prize_id]['id']
        if self.prize_id==self.masiv_prizes["row_count"]-1:
            self.bt_next.hide()

    def back_prize(self):
        self.vin_sort_text.hide()
        self.prize_id-=1
        # print(self.masiv_prizes)
        self.bt_back.show()
        self.name_prizes.setText(self.masiv_prizes["prize_data"][self.prize_id]["name"])
        image=QPixmap(self.masiv_prizes["prize_data"][self.prize_id]["image"])
        self.image_label.setPixmap(image)
        self.vin_count_text.setText(f"Количество победителей {self.masiv_prizes["prize_data"][self.prize_id]["count"]}")
        self.vin_count_text.show()

        stretch_item=self.vin_layout.takeAt(self.vin_layout.count()-1)

        if stretch_item:
            del stretch_item

        for i in range(len(self.text)):

            self.vin_layout.removeWidget(self.text[i])
            self.text[i].setParent(None)

        vin_str=""
        self.text=[]
        count=0
        self.current_index=0
        for i in self.masiv_prizes["prize_data"][self.prize_id]['vin']:
            self.text.append(QLabel(""))

            self.text[count].setFont(QtGui.QFont("Times",20))
            self.vin_layout.addWidget(self.text[count])


            # QApplication.processEvents()
            # self.unfade(self.text[i])


            count+=1
            # vin_str+=str(i)+" "
        self.start_animation()
        self.vin_layout.addStretch()

        vin_sort_str=""
        vin=self.masiv_prizes["prize_data"][self.prize_id]['vin'].copy()
        vin.sort()

        for i in vin:
            vin_sort_str+=str(i)+" "
        self.vin_sort_text.setText(vin_sort_str)

        # self.vin_text.setText(vin_str)
        self.vin_text_text.show()
        # self.vin_text.show()
        self.bt_start.hide()
        self.bt_next.show()
        self.prize_id=self.masiv_prizes["prize_data"][self.prize_id]['id']
        if self.prize_id==0:
            self.bt_back.hide()



    def load_setting(self):
        try:
            if os.path.exists("setting.json") and self.table_data.item(0,0)==None:
                with open('setting.json',"r") as f:
                    data=json.load(f)

                self.masiv_prizes=data.copy()
                self.table_data.setRowCount(len(data["prize_data"]))
                self.range_do.setValue(data["count_uchastnikov"])
                for i in range (len(data["prize_data"])):

                    self.table_data.setItem(i,0,QTableWidgetItem(data["prize_data"][i]["name"]))
                    self.table_data.setItem(i,1,QTableWidgetItem(data["prize_data"][i]["image"]))
                    self.table_data.setItem(i,2,QTableWidgetItem(str(data["prize_data"][i]["count"])))
        except:
            pass





    def form2_show(self):

        self.form.show()



    def open_file(self):
        name=QFileDialog.getOpenFileName(self,"Выбор папки","","Файлы изображений (*.png *.jpg *.bmp *.jpeg)")[0]
        self.prizes_put_edit.setText(name)

    def add_prize(self):
        name = self.prizes_name_edit.text()
        put = self.prizes_put_edit.text()
        count = self.prizes_count_edit.text()

        if name != "" and put != "":
            self.mail_text.setText("")
            self.mail_text.hide()
            row=self.table_data.rowCount()
            self.table_data.setRowCount(row+1)

            self.table_data.setItem(row,0,QTableWidgetItem(name))
            self.table_data.setItem(row,1,QTableWidgetItem(put))
            self.table_data.setItem(row,2,QTableWidgetItem(count))

            self.prizes_name_edit.setText("")
            self.prizes_put_edit.setText("")
            self.prizes_count_edit.setValue(1)

        else:
            self.mail_text.show()
            self.mail_text.setText("Не все поля заполнены!!!!")

    def save(self):
        row=self.table_data.rowCount()

        suma_prize=0
        for i in range(row):
            suma_prize+=int(self.table_data.item(i,2).text())

        if suma_prize<=int(self.range_do.text()):
            # self.mail_text.show()
            # self.mail_text.setText("1")
            row=self.table_data.rowCount()
            self.masiv_prizes["count_uchastnikov"]=self.range_do.value()
            self.masiv_prizes["row_count"]=row
            self.masiv_prizes["prize_data"]=[]

            for i in range(row):
                prize={}
                prize["id"]=i
                prize["name"]=self.table_data.item(i,0).text()
                prize["image"]=self.table_data.item(i,1).text()
                prize["count"]=int(self.table_data.item(i,2).text())
                self.masiv_prizes["prize_data"].append(prize)


            with open('setting.json', 'w') as fp:
                json.dump(self.masiv_prizes, fp)

            self.form.close()




        else:
            self.mail_text.show()
            self.mail_text.setText("Количество призов больше, чем количество участников!")



        # print(name,put,count)


    def delete_data_table(self):
        row = self.table_data.currentRow()

        if row == -1:
            msg = QMessageBox(
                parent=self,
                text="Выберите строку, которую вы хотите удалить."

            )
            msg.setWindowTitle("Window")
            msg.exec()

            return
        else:
            msg = QMessageBox.question(
                self,
                "Внимание подтвердите удаление строки!",
                "Вы действительно хотите удалить "
                f"строку <b style='color: red;'>{row+1}</b> ?"
            )
            # print(msg)
            if msg == 65536:
                return

            self.table_data.removeRow(row)


if __name__=="__main__":
    app=QApplication(sys.argv)
    ex=window()
    ex.show()
    sys.exit(app.exec())
